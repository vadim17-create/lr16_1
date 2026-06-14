from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, Sum
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.conf import settings
from decimal import Decimal
import io
from .models import Product, Category, Cart, CartItem, Order, OrderItem, Manufacturer, Profile
from .forms import RegisterForm

# --- DRF imports ---
from rest_framework import viewsets, permissions, status, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from .serializers import (
    ProductSerializer, CategorySerializer, ManufacturerSerializer,
    CartSerializer, CartItemSerializer, ProfileSerializer,
    OrderSerializer,
)

# Проверка наличия openpyxl для Excel
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def cyrillic_icontains(field, query):
    return (
        Q(**{f'{field}__contains': query}) |
        Q(**{f'{field}__contains': query.lower()}) |
        Q(**{f'{field}__contains': query.capitalize()}) |
        Q(**{f'{field}__contains': query.upper()})
    )


def _get_cart(request):
    """Получить корзину текущего пользователя из БД"""
    if not request.user.is_authenticated:
        return None
    cart, _ = Cart.objects.get_or_create(user=request.user)
    return cart


def _cart_count(request):
    """Получить количество товаров в корзине"""
    if not request.user.is_authenticated:
        return 0
    cart = _get_cart(request)
    result = cart.cartitem_set.aggregate(total=Sum('quantity'))
    return result['total'] or 0


def home(request):
    popular_products = Product.objects.order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'home.html', {
        'popular_products': popular_products,
        'categories': categories,
        'cart_count': _cart_count(request),
    })


def about(request):
    table_rows = [
        ('Удочка телескопическая', 'Лёгкая, 3 м', '35 руб.'),
        ('Катушка рыболовная', 'С фрикционом', '50 руб.'),
        ('Леска 0.25 мм', '100 м, монофильная', '12 руб.'),
        ('Поплавок 5 г', 'Плавающий', '3 руб.'),
        ('Крючки №6', 'Набор 50 шт', '7 руб.'),
        ('Силиконовые приманки', 'Набор 10 шт', '15 руб.'),
        ('Блесна вращающаяся', '7 г, цвет серебро', '8 руб.'),
        ('Сачок складной', '50 см, сетка', '25 руб.'),
        ('Эхолот рыбопоисковый', 'С ЖК-дисплеем', '120 руб.'),
        ('Воблер плавающий', '8 см, цвет рыба', '20 руб.'),
    ]
    return render(request, 'about.html', {'table_rows': table_rows, 'cart_count': _cart_count(request)})


def author(request):
    return render(request, 'author.html', {'cart_count': _cart_count(request)})


def product_list(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

    search_query = request.GET.get('search', '').strip()
    if search_query:
        products = products.filter(cyrillic_icontains('name', search_query))

    category_id = request.GET.get('category', '')
    if category_id:
        products = products.filter(category_id=category_id)

    manufacturer_id = request.GET.get('manufacturer', '')
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

    sort_by = request.GET.get('sort', '')
    if sort_by == 'cheap':
        products = products.order_by('price')
    elif sort_by == 'expensive':
        products = products.order_by('-price')

    # Пагинация — 9 товаров на странице
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'catalog.html', {
        'page_obj': page_obj,
        'products': page_obj,
        'categories': categories,
        'manufacturers': manufacturers,
        'search_query': search_query,
        'category_id': category_id,
        'manufacturer_id': manufacturer_id,
        'sort_by': sort_by,
        'total': products.count(),
        'cart_count': _cart_count(request),
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'product_detail.html', {
        'product': product,
        'cart_count': _cart_count(request),
    })


# ─── КОРЗИНА ────────────────────────────────────────────────────────────────

@login_required
@csrf_exempt
def cart_add(request, pk):
    product = get_object_or_404(Product, pk=pk)
    cart = _get_cart(request)
    qty = int(request.POST.get('quantity', 1))
    
    # Ищем существующий элемент корзины
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': qty}
    )
    if not created:
        # Если уже есть, увеличиваем количество
        cart_item.quantity = min(cart_item.quantity + qty, product.stock_quantity)
        cart_item.save()
    
    return redirect(request.META.get('HTTP_REFERER', '/catalog/'))


@login_required
@csrf_exempt
def cart_remove(request, pk):
    cart = _get_cart(request)
    CartItem.objects.filter(cart=cart, product_id=pk).delete()
    return redirect('/cart/')


@login_required
@csrf_exempt
def cart_update(request, pk):
    cart = _get_cart(request)
    product = get_object_or_404(Product, pk=pk)
    qty = int(request.POST.get('quantity', 1))
    
    if qty <= 0:
        CartItem.objects.filter(cart=cart, product=product).delete()
    else:
        CartItem.objects.update_or_create(
            cart=cart,
            product=product,
            defaults={'quantity': min(qty, product.stock_quantity)}
        )
    return redirect('/cart/')


@login_required
def cart_view(request):
    cart = _get_cart(request)
    items = cart.cartitem_set.select_related('product').all()
    total = sum(item.item_cost() for item in items)
    return render(request, 'cart.html', {
        'items': items,
        'total': total,
        'cart_count': _cart_count(request),
    })


@login_required
def cart_clear(request):
    cart = _get_cart(request)
    cart.cartitem_set.all().delete()
    return redirect('/cart/')


# ─── ОФОРМЛЕНИЕ ЗАКАЗА ──────────────────────────────────────────────────────

def _generate_excel_receipt(order):
    """Генерация чека в формате Excel"""
    if not OPENPYXL_AVAILABLE:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Чек"
    
    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ЧЕК ЗАКАЗА №{}'.format(order.id)
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
    
    # Информация о заказе
    ws['A3'] = 'Дата заказа:'
    ws['B3'] = order.created_at.strftime('%d.%m.%Y %H:%M')
    ws['A4'] = 'Покупатель:'
    ws['B4'] = order.name
    ws['A5'] = 'Телефон:'
    ws['B5'] = order.phone
    ws['A6'] = 'Адрес доставки:'
    ws['B6'] = order.address
    ws['A7'] = 'Способ оплаты:'
    ws['B7'] = dict(Order.STATUS_CHOICES).get(order.status, order.payment_method)
    
    # Таблица товаров
    ws['A9'] = '№'
    ws['B9'] = 'Товар'
    ws['C9'] = 'Кол-во'
    ws['D9'] = 'Цена (руб.)'
    ws['E9'] = 'Сумма (руб.)'
    
    for cell in ws['A9:E9']:
        cell[0].font = cell[0].font.copy(bold=True)
    
    row = 10
    for idx, item in enumerate(order.items.all(), 1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = item.product.name
        ws[f'C{row}'] = item.quantity
        ws[f'D{row}'] = float(item.price)
        ws[f'E{row}'] = float(item.item_cost())
        row += 1
    
    # Итого
    ws[f'A{row}'] = 'ИТОГО:'
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
    ws[f'D{row}'] = float(order.total_amount)
    ws[f'D{row}'].font = ws[f'D{row}'].font.copy(bold=True)
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Сохранение в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _send_order_email(order, user_email):
    """Отправка чека по электронной почте"""
    if not OPENPYXL_AVAILABLE:
        print("❌ openpyxl не установлен")
        return False

    excel_file = _generate_excel_receipt(order)
    if not excel_file:
        print("❌ Не удалось создать Excel файл")
        return False

    print(f"📧 Отправка email на {user_email}...")
    
    email = EmailMessage(
        subject=f'Заказ №{order.id} — SALMO Shop',
        body=f'Здравствуйте, {order.name}!\n\n'
             f'Благодарим за заказ в SALMO Shop!\n\n'
             f'Номер заказа: {order.id}\n'
             f'Сумма заказа: {order.total_amount} руб.\n'
             f'Адрес доставки: {order.address}\n\n'
             f'Чек прикреплён к этому письму.',
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[user_email],
    )
    email.attach(f'receipt_order_{order.id}.xlsx', excel_file.getvalue(),
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    try:
        email.send(fail_silently=False)
        print(f"✅ Email успешно отправлен на {user_email}")
        return True
    except Exception as e:
        print(f"❌ Ошибка отправки email: {e}")
        import traceback
        traceback.print_exc()
        return False


@login_required
def checkout(request):
    cart = _get_cart(request)
    if not cart or not cart.cartitem_set.exists():
        return redirect('/cart/')

    items = cart.cartitem_set.select_related('product').all()
    total = sum(item.item_cost() for item in items)
    error = ''
    form_data = {'name': '', 'phone': '', 'email': '', 'address': ''}

    if request.method == 'POST':
        form_data = {
            'name': request.POST.get('name', '').strip(),
            'phone': request.POST.get('phone', '').strip(),
            'email': request.POST.get('email', '').strip(),
            'address': request.POST.get('address', '').strip(),
        }
        payment_method = request.POST.get('payment', 'card')
        
        if not all(form_data.values()):
            error = '⚠️ Заполните все поля.'
        else:
            # Создание заказа
            order = Order.objects.create(
                user=request.user,
                name=form_data['name'],
                phone=form_data['phone'],
                address=form_data['address'],
                payment_method=payment_method,
                total_amount=total,
                email=form_data['email'],
            )
            
            # Перенос элементов корзины в заказ
            for item in items:
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    quantity=item.quantity,
                    price=item.product.price,
                )
                # Уменьшаем количество на складе
                item.product.stock_quantity -= item.quantity
                item.product.save()
            
            # Очистка корзины
            cart.cartitem_set.all().delete()
            
            # Отправка email с чеком на введённый email
            _send_order_email(order, form_data['email'])
            
            return redirect('/order-success/')

    return render(request, 'checkout.html', {
        'items': items,
        'total': total,
        'error': error,
        'form_data': form_data,
        'cart_count': _cart_count(request),
    })


def order_success(request):
    return render(request, 'order_success.html', {'cart_count': 0})


# ─── РЕГИСТРАЦИЯ ────────────────────────────────────────────────────────────

def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Профиль создаётся автоматически через signal
            # Создаём пустую корзину
            Cart.objects.get_or_create(user=user)
            login(request, user)
            return redirect('/')
        else:
            print("Ошибки формы:", form.errors)
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})


# ─── DRF API ViewSets ────────────────────────────────────────────────────────

class IsAdminOrReadOnly(permissions.BasePermission):
    """GET доступен всем, POST/PUT/PATCH/DELETE — только администраторам"""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        if not request.user.is_authenticated:
            return False
        return hasattr(request.user, 'profile') and request.user.profile.is_admin


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAdminOrReadOnly]


class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]


class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Покупатель видит только свои заказы, админ — все"""
        user = self.request.user
        if hasattr(user, 'profile') and user.profile.is_admin:
            return Order.objects.all()
        return Order.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ─── /api/me/ endpoint ──────────────────────────────────────────────────────

@api_view(['GET', 'PATCH'])
@permission_classes([permissions.IsAuthenticated])
def me_view(request):
    """GET /api/me/ — профиль текущего пользователя"""
    """PATCH /api/me/ — изменение профиля"""
    profile, _ = Profile.objects.get_or_create(user=request.user)

    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)

    elif request.method == 'PATCH':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ─── ЛИЧНЫЙ КАБИНЕТ ───────────────────────────────────────────────────────

@login_required
def profile_view(request):
    """Страница личного кабинета"""
    profile, _ = Profile.objects.get_or_create(user=request.user)
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    categories = Category.objects.all()

    if request.method == 'POST':
        profile.full_name = request.POST.get('full_name', '').strip()
        profile.phone = request.POST.get('phone', '').strip()
        profile.address = request.POST.get('address', '').strip()
        profile.city = request.POST.get('city', '').strip()
        profile.postal_code = request.POST.get('postal_code', '').strip()
        fav_cat = request.POST.get('favorite_category', '')
        if fav_cat:
            profile.favorite_category_id = fav_cat
        else:
            profile.favorite_category = None
        profile.save()
        return redirect('/profile/')

    return render(request, 'profile.html', {
        'profile': profile,
        'orders': orders,
        'categories': categories,
        'cart_count': _cart_count(request),
    })


@login_required
def order_detail_view(request, pk):
    """Детали заказа"""
    order = get_object_or_404(Order, pk=pk)
    # Покупатель может смотреть только свои заказы, админ — все
    if not (hasattr(request.user, 'profile') and request.user.profile.is_admin):
        if order.user != request.user:
            return redirect('/profile/')
    return render(request, 'order_detail.html', {
        'order': order,
        'cart_count': _cart_count(request),
    })


@login_required
def settings_view(request):
    """Страница настроек — смена пароля"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect('/profile/')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'settings.html', {
        'form': form,
        'cart_count': _cart_count(request),
    })
